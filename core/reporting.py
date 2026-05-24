from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Dict, Mapping

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


REVIEW_STATUSES = {"REVIEW_SUBSTITUTE", "REVIEW_ALTERNATIVE", "UOM_REVIEW"}
NO_MATCH_STATUSES = {"NO_MATCH", "HIGHER_PRICE", "NO_SAVINGS"}


def summarize_results(results: pd.DataFrame) -> Dict[str, float | int]:
    old_spend = float(results["old_spend"].fillna(0).sum()) if "old_spend" in results else 0.0
    matched_spend = float(results.loc[~results["match_status"].eq("NO_MATCH"), "old_spend"].fillna(0).sum())
    confirmed = float(results.loc[results["savings_bucket"].eq("confirmed_savings"), "estimated_savings"].fillna(0).sum())
    potential = float(
        results.loc[results["savings_bucket"].eq("potential_review_savings"), "estimated_savings"].fillna(0).sum()
    )
    auto_confirmed_rows = int(results["match_status"].eq("AUTO_CONFIRMED").sum())
    review_rows = int(results["match_status"].isin(REVIEW_STATUSES).sum())
    no_match_or_higher = int(results["match_status"].isin(NO_MATCH_STATUSES).sum())
    return {
        "total_old_spend_analyzed": old_spend,
        "matched_spend": matched_spend,
        "confirmed_savings": confirmed,
        "potential_review_savings": potential,
        "confirmed_savings_percent": confirmed / old_spend if old_spend else 0,
        "rows_auto_confirmed": auto_confirmed_rows,
        "rows_needing_review": review_rows,
        "no_match_higher_price_rows": no_match_or_higher,
    }


def summarize_coverage(results: pd.DataFrame) -> Dict[str, float | int]:
    rows_analyzed = int(len(results))
    rows_with_candidate = int(results["sourceclub_product_id"].astype("string").fillna("").ne("").sum())
    rows_auto_confirmed = int(results["match_status"].eq("AUTO_CONFIRMED").sum())
    rows_needing_review = int(results["match_status"].isin(REVIEW_STATUSES).sum())
    rows_no_match_higher = int(results["match_status"].isin(NO_MATCH_STATUSES).sum())

    return {
        "rows_analyzed": rows_analyzed,
        "rows_with_any_candidate_match": rows_with_candidate,
        "rows_auto_confirmed": rows_auto_confirmed,
        "rows_needing_review": rows_needing_review,
        "rows_no_match_higher_price": rows_no_match_higher,
        "catalog_coverage_percent": rows_with_candidate / rows_analyzed if rows_analyzed else 0,
    }


def coverage_label(coverage_percent: float) -> str:
    if coverage_percent >= 0.70:
        return "Strong coverage"
    if coverage_percent >= 0.30:
        return "Partial coverage"
    return "Low coverage"


def summary_frame(
    summary: Dict[str, float | int],
    coverage: Dict[str, float | int] | None = None,
    metadata: Mapping[str, object] | None = None,
) -> pd.DataFrame:
    labels = {
        "total_old_spend_analyzed": "Total old spend analyzed",
        "matched_spend": "Matched spend",
        "confirmed_savings": "Confirmed savings",
        "potential_review_savings": "Potential/review savings",
        "confirmed_savings_percent": "Confirmed savings %",
        "rows_auto_confirmed": "Rows auto-confirmed",
        "rows_needing_review": "Rows needing review",
        "no_match_higher_price_rows": "No-match / higher-price rows",
    }
    rows = [{"Summary item": labels.get(key, key), "Value": value} for key, value in summary.items()]

    if coverage:
        rows.extend(
            [
                {"Summary item": "Rows analyzed", "Value": coverage["rows_analyzed"]},
                {"Summary item": "Rows with candidate match", "Value": coverage["rows_with_any_candidate_match"]},
                {"Summary item": "Rows auto-confirmed", "Value": coverage["rows_auto_confirmed"]},
                {"Summary item": "Rows needing review", "Value": coverage["rows_needing_review"]},
                {"Summary item": "No-match/higher-price rows", "Value": coverage["rows_no_match_higher_price"]},
                {"Summary item": "Catalog coverage %", "Value": coverage["catalog_coverage_percent"]},
                {"Summary item": "Coverage label", "Value": coverage_label(float(coverage["catalog_coverage_percent"]))},
            ]
        )

    if metadata:
        rows.extend([{"Summary item": key, "Value": value} for key, value in metadata.items()])

    return pd.DataFrame(rows)


def split_result_tables(results: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    review_queue = results[results["match_status"].isin(REVIEW_STATUSES)].copy()
    no_match = results[
        results["match_status"].isin(NO_MATCH_STATUSES) | results["savings_bucket"].eq("excluded_or_no_savings")
    ].copy()
    candidate_rows = results[results["sourceclub_product_id"].astype("string").fillna("").ne("")].copy()
    library_updates = candidate_rows[
        [
            "customer_product_name",
            "current_supplier",
            "distributor_item_number",
            "manufacturer",
            "manufacturer_item_number",
            "sourceclub_product_id",
            "best_supplier",
            "best_supplier_product_sku",
            "best_supplier_product_name",
            "match_status",
            "match_type",
            "similarity_score",
        ]
    ].copy()
    library_updates["training_status"] = "candidate_for_review"
    library_updates["reviewer_decision"] = ""
    return {
        "review_queue": review_queue,
        "no_match_higher_price": no_match,
        "match_library_updates": library_updates,
    }


def make_excel_workbook(
    results: pd.DataFrame,
    summary: Dict[str, float | int],
    coverage: Dict[str, float | int] | None = None,
    metadata: Mapping[str, object] | None = None,
) -> bytes:
    tables = split_result_tables(results)
    workbook_metadata = {
        "Demo mode": "",
        "Source file name": "",
        "Catalog source": "",
        "Generated timestamp": datetime.now().isoformat(timespec="seconds"),
        "Synthetic data warning": (
            "Synthetic demo data only. Not real customer, PHI, proprietary supplier, or SourceClub pricing data."
        ),
    }
    if metadata:
        workbook_metadata.update(metadata)

    savings_columns = [
        "customer_product_name",
        "current_supplier",
        "distributor_item_number",
        "manufacturer",
        "manufacturer_item_number",
        "quantity",
        "uom",
        "old_spend",
        "current_unit_price",
        "best_supplier",
        "best_supplier_product_sku",
        "best_supplier_product_name",
        "best_price",
        "match_status",
        "match_type",
        "similarity_score",
        "sourceclub_new_spend",
        "estimated_savings",
        "estimated_savings_percent",
        "savings_bucket",
        "match_notes",
    ]
    savings_analysis = results[[column for column in savings_columns if column in results.columns]].copy()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_frame(summary, coverage, workbook_metadata).to_excel(writer, index=False, sheet_name="Summary")
        savings_analysis.to_excel(writer, index=False, sheet_name="Savings Analysis")
        tables["review_queue"].to_excel(writer, index=False, sheet_name="Review Queue")
        tables["no_match_higher_price"].to_excel(writer, index=False, sheet_name="No Match Higher Price")
        tables["match_library_updates"].to_excel(writer, index=False, sheet_name="Match Library Updates")

        header_fill = PatternFill("solid", fgColor="E7F0ED")
        header_font = Font(bold=True, color="1F2933")
        summary_label_font = Font(bold=True, color="1F2933")

        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            for header_cell in sheet[1]:
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = Alignment(vertical="center")
            for column_cells in sheet.columns:
                header = str(column_cells[0].value or "")
                width = min(max(len(header) + 2, 12), 42)
                sheet.column_dimensions[column_cells[0].column_letter].width = width

        summary_sheet = writer.book["Summary"]
        for row in summary_sheet.iter_rows(min_row=2, max_col=2):
            row[0].font = summary_label_font
            row[0].alignment = Alignment(vertical="top")
            row[1].alignment = Alignment(vertical="top", wrap_text=True)
            label = str(row[0].value or "")
            if "%" in label and isinstance(row[1].value, (float, int)):
                row[1].number_format = "0.0%"
            elif "savings" in label.lower() or "spend" in label.lower():
                row[1].number_format = "$#,##0.00"
        summary_sheet.column_dimensions["A"].width = 34
        summary_sheet.column_dimensions["B"].width = 72

    return output.getvalue()
